"""
API Views for CPSU Virtual Health Assistant
Implements all endpoints for student and clinic staff
"""

from rest_framework import viewsets, status, generics, serializers
from rest_framework.decorators import action, api_view, permission_classes, renderer_classes
from rest_framework.response import Response
from rest_framework.exceptions import APIException
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.authtoken.models import Token
from django.contrib.auth import authenticate, get_user_model
from django.utils import timezone
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Count, Q
from django.db import IntegrityError
from datetime import timedelta
import uuid
import logging

from .models import SymptomRecord, HealthInsight, ChatSession, ConsentLog, AuditLog, DepartmentStats, EmergencyAlert, Medication, MedicationLog, FollowUp, Message, Appointment
from .serializers import (
    UserRegistrationSerializer, UserProfileSerializer,
    SymptomRecordSerializer, SymptomSubmissionSerializer,
    DiseasePredictionSerializer, HealthInsightSerializer,
    ChatSessionSerializer, ChatMessageSerializer,
    ConsentLogSerializer, AuditLogSerializer,
    DepartmentStatsSerializer, DashboardStatsSerializer,
    EmergencyAlertSerializer, EmergencyTriggerSerializer,
    MedicationSerializer, MedicationCreateSerializer, MedicationLogSerializer,
    FollowUpSerializer, FollowUpResponseSerializer,
    MessageSerializer, AppointmentSerializer
)
from .permissions import IsStudent, IsClinicStaff, IsOwnerOrStaff, CanModifyProfile, HasDataConsent
from .ml_service import get_ml_predictor

logger = logging.getLogger(__name__)
from .llm_service import AIInsightGenerator

# ------------------------------------------------------------------
# Rasa integration – commented out until Rasa is deployed.
# Uncomment the two lines below and the Rasa flow inside
# send_chat_message() when you're ready to use Rasa.
# ------------------------------------------------------------------
# from .rasa_service import RasaChatService
# rasa_service = RasaChatService()

# Get singleton instances
ml_predictor = get_ml_predictor()
ai_generator = AIInsightGenerator()

User = get_user_model()


# ============================================================================
# Admin Account Management (superuser-only)
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def admin_list_accounts(request):
    """
    List all user accounts (superuser only)
    GET /api/admin/accounts/
    Query params: role, search, is_active
    """
    if not request.user.is_superuser:
        return Response({'error': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)

    users = User.objects.all().order_by('-date_joined')

    # Filters
    role = request.GET.get('role')
    if role:
        users = users.filter(role=role)

    search = request.GET.get('search', '').strip()
    if search:
        users = users.filter(
            Q(school_id__icontains=search) |
            Q(name__icontains=search)
        )

    is_active = request.GET.get('is_active')
    if is_active is not None:
        users = users.filter(is_active=is_active.lower() == 'true')

    data = []
    for u in users[:200]:
        data.append({
            'id': str(u.id) if hasattr(u, 'id') else str(u.pk),
            'school_id': u.school_id,
            'name': u.name,
            'role': u.role,
            'department': u.department,
            'is_active': u.is_active,
            'is_superuser': u.is_superuser,
            'date_joined': u.date_joined.isoformat(),
        })

    return Response({
        'count': len(data),
        'accounts': data
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_create_account(request):
    """
    Create a staff or student account (superuser only)
    POST /api/admin/accounts/create/
    Body: { school_id, password, name, role, department }
    """
    if not request.user.is_superuser:
        return Response({'error': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)

    school_id = request.data.get('school_id', '').strip()
    password = request.data.get('password', '').strip()
    name = request.data.get('name', '').strip()
    role = request.data.get('role', 'staff')
    department = request.data.get('department', '')

    if not school_id or not password:
        return Response({'error': 'school_id and password are required'}, status=status.HTTP_400_BAD_REQUEST)

    if len(password) < 6:
        return Response({'error': 'Password must be at least 6 characters'}, status=status.HTTP_400_BAD_REQUEST)

    if User.objects.filter(school_id=school_id).exists():
        return Response({'error': f'User with school_id "{school_id}" already exists'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        user = User.objects.create_user(
            school_id=school_id,
            password=password,
            name=name,
            role=role,
            department=department,
            is_staff=role == 'staff',
        )

        return Response({
            'message': f'{role.capitalize()} account created successfully',
            'account': {
                'id': str(user.id) if hasattr(user, 'id') else str(user.pk),
                'school_id': user.school_id,
                'name': user.name,
                'role': user.role,
                'is_active': user.is_active,
            }
        }, status=status.HTTP_201_CREATED)
    except Exception as e:
        logger.error(f"Error creating account: {e}")
        return Response({'error': 'Failed to create account'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def admin_toggle_account(request, school_id):
    """
    Toggle account active status (superuser only)
    PATCH /api/admin/accounts/<school_id>/toggle/
    """
    if not request.user.is_superuser:
        return Response({'error': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)

    try:
        user = User.objects.get(school_id=school_id)
    except User.DoesNotExist:
        return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

    if user == request.user:
        return Response({'error': 'Cannot deactivate your own account'}, status=status.HTTP_400_BAD_REQUEST)

    user.is_active = not user.is_active
    user.save()

    return Response({
        'message': f'Account {"activated" if user.is_active else "deactivated"}',
        'is_active': user.is_active
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def admin_reset_password(request, school_id):
    """
    Reset a user's password (superuser only)
    POST /api/admin/accounts/<school_id>/reset-password/
    Body: { new_password }
    """
    if not request.user.is_superuser:
        return Response({'error': 'Admin access required'}, status=status.HTTP_403_FORBIDDEN)

    try:
        user = User.objects.get(school_id=school_id)
    except User.DoesNotExist:
        return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

    new_password = request.data.get('new_password', '').strip()
    if not new_password or len(new_password) < 6:
        return Response({'error': 'Password must be at least 6 characters'}, status=status.HTTP_400_BAD_REQUEST)

    user.set_password(new_password)
    user.save()

    return Response({'message': f'Password reset for {user.school_id}'})


# ============================================================================
# Authentication Views
# ============================================================================

from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.core.mail import send_mail
from django.conf import settings

@api_view(['POST'])
@permission_classes([AllowAny])
def register_user(request):
    """
    Register new user (student or staff)
    POST /api/auth/register/
    """
    serializer = UserRegistrationSerializer(data=request.data)
    
    if serializer.is_valid():
        # Determine role based on registration context
        # Restrict public registration to students only. Staff must be created by admin.
        role = 'student'
        
        user = serializer.save(role=role)
        
        # Create auth token
        token, created = Token.objects.get_or_create(user=user)
        
        # Log consent if given
        if user.data_consent_given:
            ConsentLog.objects.create(
                user=user,
                action='granted',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
            )
        
        return Response({
            'token': token.key,
            'user': UserProfileSerializer(user).data,
            'message': 'Registration successful'
        }, status=status.HTTP_201_CREATED)
    
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([AllowAny])
def login_user(request):
    """
    User login with school_id and password
    POST /api/auth/login/
    """
    school_id = request.data.get('school_id')
    password = request.data.get('password')
    
    if not school_id or not password:
        return Response(
            {'error': 'School ID and password are required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    user = authenticate(username=school_id, password=password)
    
    if user:
        token, created = Token.objects.get_or_create(user=user)
        
        return Response({
            'token': token.key,
            'user': UserProfileSerializer(user).data,
            'message': 'Login successful'
        })
    
    return Response(
        {'error': 'Invalid credentials'},
        status=status.HTTP_401_UNAUTHORIZED
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_user(request):
    """
    User logout (delete auth token)
    POST /api/auth/logout/
    """
    try:
        request.user.auth_token.delete()
    except (AttributeError, ObjectDoesNotExist):
        pass
    return Response({'message': 'Logout successful'})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def change_password(request):
    """
    Change the current user's password.
    POST /api/auth/change-password/
    """
    user = request.user
    old_password = request.data.get('old_password')
    new_password = request.data.get('new_password')
    
    if not old_password or not new_password:
        return Response({'error': 'Please provide both current and new passwords.'}, status=status.HTTP_400_BAD_REQUEST)
        
    if not user.check_password(old_password):
        return Response({'error': 'Incorrect current password.'}, status=status.HTTP_400_BAD_REQUEST)
        
    if len(new_password) < 6:
        return Response({'error': 'New password must be at least 6 characters long.'}, status=status.HTTP_400_BAD_REQUEST)
        
    user.set_password(new_password)
    user.save()
    
    # Optional: Delete current token and issue new one for security
    user.auth_token.delete()
    new_token, _ = Token.objects.get_or_create(user=user)
    
    AuditLog.objects.create(
        user=user,
        action="PASSWORD_CHANGE",
        ip_address=request.META.get('REMOTE_ADDR'),
        details={"source": "user_profile"}
    )
    
    return Response({
        'message': 'Password changed successfully.',
        'token': new_token.key
    })


@api_view(['POST'])
@permission_classes([AllowAny])
def forgot_password(request):
    """
    Send a password reset email via Brevo.
    POST /api/auth/forgot-password/
    """
    identifier = request.data.get('identifier')
    if not identifier:
        return Response({'error': 'Please provide your School ID or Email address.'}, status=status.HTTP_400_BAD_REQUEST)
        
    user = User.objects.filter(Q(email__iexact=identifier) | Q(school_id__iexact=identifier)).first()
    
    if user and user.email:
        # Generate reset token and uid
        token = default_token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        
        # Build reset link pointing to Vue frontend
        frontend_url = request.META.get('HTTP_ORIGIN', 'http://localhost:5173')
        reset_link = f"{frontend_url}/reset-password?uid={uid}&token={token}"
        
        subject = "Action Required: Password Reset for CPSU Health Assistant"
        body = f"""Hello {user.name},

You or someone else has requested a password reset for your CPSU Virtual Health Assistant account.

Please click the link below to reset your password:
{reset_link}

If you did not request this, please ignore this email. Your password will remain unchanged.

Best regards,
The CPSU Clinic Team"""
        
        try:
            send_mail(
                subject,
                body,
                settings.DEFAULT_FROM_EMAIL,
                [user.email],
                fail_silently=False,
            )
            logger.info(f"Password reset email sent to {user.email}")
        except Exception as e:
            logger.error(f"Failed to send reset email: {e}")
            return Response(
                {'error': 'Failed to send email. Ensure your email address is valid.'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
            
    # Always return success message to prevent user enumeration
    return Response({'message': 'If an account exists with that email address, a password reset link has been sent.'})


@api_view(['POST'])
@permission_classes([AllowAny])
def reset_password_confirm(request):
    """
    Verify reset token and set new password.
    POST /api/auth/reset-password-confirm/
    """
    uidb64 = request.data.get('uid')
    token = request.data.get('token')
    new_password = request.data.get('new_password')
    
    if not uidb64 or not token or not new_password:
        return Response({'error': 'Missing required fields.'}, status=status.HTTP_400_BAD_REQUEST)
        
    if len(new_password) < 6:
        return Response({'error': 'Password must be at least 6 characters long.'}, status=status.HTTP_400_BAD_REQUEST)
        
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
        
    if user is not None and default_token_generator.check_token(user, token):
        user.set_password(new_password)
        user.save()
        
        # Invalidate old auth token to force logout elsewhere
        try:
            user.auth_token.delete()
        except ObjectDoesNotExist:
            pass
            
        AuditLog.objects.create(
            user=user,
            action="PASSWORD_RESET",
            ip_address=request.META.get('REMOTE_ADDR'),
            details={"source": "email_token"}
        )
        return Response({'message': 'Password has been reset successfully. You can now log in.'})
    else:
        return Response({'error': 'The password reset link is invalid or has expired.'}, status=status.HTTP_400_BAD_REQUEST)


# ============================================================================
# User Profile Views
# ============================================================================

class UserProfileView(generics.RetrieveUpdateAPIView):
    """
    Get or update user profile
    GET/PUT/PATCH /api/profile/
    """
    serializer_class = UserProfileSerializer
    permission_classes = [IsAuthenticated, CanModifyProfile]
    
    def get_object(self):
        return self.request.user
    
    def update(self, request, *args, **kwargs):
        # Check for immutable fields
        immutable_fields = {'school_id', 'role'}
        if any(field in request.data for field in immutable_fields):
            return Response(
                {'error': 'Cannot modify school_id or role'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return super().update(request, *args, **kwargs)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def update_consent(request):
    """
    Update data consent status
    POST /api/profile/consent/
    """
    consent_given = request.data.get('data_consent_given')
    
    if consent_given is None:
        return Response(
            {'error': 'data_consent_given field is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    user = request.user
    previous_consent = user.data_consent_given
    user.data_consent_given = consent_given
    
    if consent_given and not user.consent_date:
        user.consent_date = timezone.now()
    
    user.save()
    
    # Log consent change
    action = 'granted' if consent_given else 'revoked'
    if previous_consent == consent_given:
        action = 'updated'
    
    ConsentLog.objects.create(
        user=user,
        action=action,
        ip_address=request.META.get('REMOTE_ADDR'),
        user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
    )
    
    return Response({
        'message': f'Consent {action}',
        'data_consent_given': user.data_consent_given,
        'consent_date': user.consent_date
    })


# ============================================================================
# Symptom & ML Views
# ============================================================================

class SymptomRecordViewSet(viewsets.ModelViewSet):
    """
    Symptom record CRUD operations
    Students can only access their own records
    Staff can access all records
    """
    serializer_class = SymptomRecordSerializer
    permission_classes = [IsAuthenticated, IsOwnerOrStaff]

    @action(detail=True, methods=['patch'], permission_classes=[IsAuthenticated, IsClinicStaff])
    def diagnosis(self, request, pk=None):
        """
        PATCH /api/symptoms/{id}/diagnosis/
        Staff-only: set the final diagnosis for a symptom record.
        """
        record = self.get_object()
        staff_diag = request.data.get('staff_diagnosis', '').strip()
        if not staff_diag:
            return Response({'error': 'staff_diagnosis is required.'}, status=status.HTTP_400_BAD_REQUEST)
        record.staff_diagnosis = staff_diag
        record.final_diagnosis = staff_diag
        record.save(update_fields=['staff_diagnosis', 'final_diagnosis', 'updated_at'])
        return Response({
            'id': str(record.id),
            'staff_diagnosis': record.staff_diagnosis,
            'final_diagnosis': record.final_diagnosis,
        })

    def get_queryset(self):
        user = self.request.user
        
        if user.role == 'staff':
            # Staff can see all records
            queryset = SymptomRecord.objects.all().select_related('student')
        else:
            # Students see only their own
            queryset = SymptomRecord.objects.filter(student=user)
        
        # Filter by date range if provided
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        
        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__lte=end_date)
        
        return queryset.order_by('-created_at')


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsStudent, HasDataConsent])
def submit_symptoms(request):
    """
    Submit symptoms and get disease prediction
    POST /api/symptoms/submit/
    """
    serializer = SymptomSubmissionSerializer(data=request.data)
    
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    data = serializer.validated_data
    
    try:
        # Get ML prediction
        predictor = get_ml_predictor()
        prediction_result = predictor.predict(data['symptoms'])
        
        # Create symptom record
        record = SymptomRecord.objects.create(
            student=request.user,
            symptoms=data['symptoms'],
            duration_days=data['duration_days'],
            severity=data['severity'],
            on_medication=data.get('on_medication', False),
            medication_adherence=data.get('medication_adherence'),
            patient_age=data.get('patient_age'),
            patient_sex=data.get('patient_sex', ''),
            predicted_disease=prediction_result['predicted_disease'],
            confidence_score=prediction_result['confidence_score'],
            top_predictions=prediction_result['top_predictions'],
            is_communicable=prediction_result['is_communicable'],
            is_acute=prediction_result['is_acute'],
            icd10_code=prediction_result['icd10_code']
        )
        
        # Check referral criteria
        record.check_referral_criteria()
        record.save()
        
        # Auto-create follow-up (3 days from now)
        followup = FollowUp.create_from_symptom(record, days_ahead=3)
        
        # Prepare response
        response_data = {
            'record_id': str(record.id),
            'prediction': prediction_result,
            'requires_referral': record.requires_referral,
            'referral_message': 'You have reported symptoms 5+ times in the past 30 days. Please visit the clinic for evaluation.' if record.requires_referral else None,
            'followup_scheduled': followup.scheduled_date.isoformat()
        }
        
        return Response(response_data, status=status.HTTP_201_CREATED)
    
    except Exception as e:
        return Response(
            {'error': f'Prediction failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_available_symptoms(request):
    """
    Get list of all symptoms the ML model recognizes
    GET /api/symptoms/available/
    """
    try:
        predictor = get_ml_predictor()
        symptoms = predictor.get_available_symptoms()
        
        return Response({
            'count': len(symptoms),
            'symptoms': sorted(symptoms)
        })
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# ============================================================================
# AI Chat & Insights Views
# ============================================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated, IsStudent, HasDataConsent])
def start_chat_session(request):
    """
    Start a new AI chat session
    POST /api/chat/start/
    """
    language = request.data.get('language', 'english')
    
    session = ChatSession.objects.create(
        student=request.user,
        language=language
    )
    
    return Response({
        'session_id': str(session.id),
        'message': 'Chat session started',
        'language': language
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsStudent, HasDataConsent])
def send_chat_message(request):
    """
    Send message in AI chat – **Cohere-powered (Rasa bypassed)**.
    Cohere handles BOTH symptom extraction AND disease prediction.
    No ML model is used here (the Symptom Checker page uses the ML model).
    POST /api/chat/message/

    When Rasa is deployed, uncomment the RASA FLOW block below and
    comment out the COHERE-ONLY FLOW block.
    """
    serializer = ChatMessageSerializer(data=request.data)

    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    data = serializer.validated_data
    message = data['message']
    language = data.get('language', 'english')
    session_id = data.get('session_id')

    try:
        if session_id:
            session = ChatSession.objects.get(id=session_id, student=request.user)
        else:
            return Response(
                {'error': 'session_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # =================================================================
        # COHERE-ONLY FLOW  (active while Rasa is not deployed)
        #
        # Step 1: Cohere extracts symptoms AND predicts disease in one call
        # Step 2: Cohere generates a friendly diagnostic response
        # Step 3: Save SymptomRecord to DB (shows in history, same as Rasa)
        # =================================================================

        # ── Follow-up question state (stored in session.topics_discussed) ──────
        topics = session.topics_discussed or []
        # topics is a list; we store our state object as the last element when needed
        followup_state = None
        if topics and isinstance(topics[-1], dict) and topics[-1].get('_followup_pending'):
            followup_state = topics[-1]

        # ── Step 1: If we are waiting for the student's follow-up answer ─────────
        if followup_state:
            # Student replied to our clarifying questions — build enriched context
            original_message = followup_state.get('original_message', '')
            original_symptoms = followup_state.get('symptoms', [])
            symptoms_str = ', '.join(s.replace('_', ' ') for s in original_symptoms)
            enriched_message = (
                f"Original complaint: {original_message}. "
                f"Symptoms identified: {symptoms_str}. "
                f"Student's follow-up answer: {message}"
            )
            diagnosis = ai_generator.extract_and_predict(enriched_message)
            # Preserve original symptoms if Cohere didn't re-extract them
            if not diagnosis.get('extracted_symptoms'):
                diagnosis['extracted_symptoms'] = original_symptoms
            if not diagnosis.get('has_symptoms'):
                diagnosis['has_symptoms'] = bool(original_symptoms)
            # Clear follow-up state from session
            session.topics_discussed = [t for t in topics if not (isinstance(t, dict) and t.get('_followup_pending'))]
            session.save(update_fields=['topics_discussed'])
            has_symptoms = diagnosis.get('has_symptoms', True)
            logger.info(f"Follow-up answer received — enriched diagnosis: {diagnosis.get('predicted_disease')}")

        else:
            # ── Step 1 (normal): Cohere extracts symptoms + predicts disease ────
            diagnosis = ai_generator.extract_and_predict(message)
            has_symptoms = diagnosis.get('has_symptoms', False)
            logger.info(
                f"Cohere diagnosis: has_symptoms={has_symptoms}, "
                f"symptoms={diagnosis.get('extracted_symptoms')}, "
                f"disease={diagnosis.get('predicted_disease')}"
            )

            # ── Ask follow-up questions on FIRST symptom detection ───────────
            if has_symptoms and diagnosis.get('extracted_symptoms'):
                try:
                    followup_response = ai_generator.generate_followup_questions(
                        diagnosis['extracted_symptoms'], message
                    )
                except Exception as fq_err:
                    logger.error(f"Follow-up question generation failed: {fq_err}")
                    followup_response = None

                if followup_response:
                    # Save state so next message knows to skip straight to diagnosis
                    new_state = {
                        '_followup_pending': True,
                        'original_message': message,
                        'symptoms': diagnosis['extracted_symptoms'],
                    }
                    session.topics_discussed = [t for t in topics
                                                if not (isinstance(t, dict) and t.get('_followup_pending'))]
                    session.topics_discussed.append(new_state)
                    session.save(update_fields=['topics_discussed'])
                    return Response({
                        'response': followup_response,
                        'session_id': str(session.id),
                        'awaiting_followup': True,
                        'diagnosis_saved': False,
                    })

        # ── Step 2 – Generate full diagnostic chat response ───────────────────
        if has_symptoms and diagnosis.get('predicted_disease'):
            try:
                response_text = ai_generator.generate_diagnosis_response(message, diagnosis)
            except Exception as resp_err:
                logger.error(f"Diagnosis response generation failed: {resp_err}")
                response_text = ai_generator.generate_chat_response(
                    message=message,
                    context={'language': language}
                )
        else:
            try:
                response_text = ai_generator.generate_chat_response(
                    message=message,
                    context={'language': language, 'session_id': str(session_id)}
                )
            except Exception as chat_err:
                logger.error(f"Chat response generation failed: {chat_err}")
                response_text = (
                    "Thank you for your message. I'm here to help with health concerns. "
                    "Could you describe your symptoms so I can assist you better?"
                )

        if not response_text or not response_text.strip():
            response_text = (
                "Thank you for your message. I'm experiencing technical difficulties. "
                "Please consult with our clinic staff for proper evaluation."
            )

        # Step 3 – Save SymptomRecord (same shape as old Rasa flow)
        record_id = None
        if has_symptoms and diagnosis.get('predicted_disease'):
            try:
                severity_map = {'mild': 1, 'moderate': 2, 'severe': 3}
                severity_value = diagnosis.get('severity', 'moderate')
                if isinstance(severity_value, str):
                    severity_int = severity_map.get(severity_value.lower(), 2)
                else:
                    severity_int = int(severity_value) if severity_value else 2

                record = SymptomRecord.objects.create(
                    student=request.user,
                    symptoms=diagnosis['extracted_symptoms'],
                    duration_days=diagnosis.get('duration_days', 1),
                    severity=severity_int,
                    predicted_disease=diagnosis['predicted_disease'],
                    confidence_score=diagnosis.get('confidence_score', 0.0),
                    top_predictions=diagnosis.get('top_predictions', []),
                    is_communicable=diagnosis.get('is_communicable', False),
                    is_acute=diagnosis.get('is_acute', True),
                    icd10_code=diagnosis.get('icd10_code', ''),
                )

                record.check_referral_criteria()
                record.save()

                FollowUp.create_from_symptom(record, days_ahead=3)

                record_id = str(record.id)
                logger.info(f"Created symptom record {record_id} from Cohere chat diagnosis")

            except Exception as e:
                logger.error(f"Failed to create symptom record from chat: {e}")

        return Response({
            'response': response_text,
            'session_id': str(session_id),
            'source': 'cohere',
            'buttons': [],
            'rasa_available': False,
            'record_id': record_id,
            'diagnosis_saved': record_id is not None,
        })

        # =================================================================
        # RASA FLOW  (uncomment when Rasa is deployed)
        #
        # Also uncomment the rasa_service import + singleton at the top of
        # this file.  Then comment out the entire COHERE-ONLY FLOW above.
        # =================================================================
        #
        # rasa_response = rasa_service.send_message(
        #     message=message,
        #     sender_id=str(session_id),
        #     metadata={
        #         'language': language,
        #         'user_id': str(request.user.id),
        #         'django_api': request.build_absolute_uri('/api/')
        #     }
        # )
        #
        # if rasa_service.should_use_llm_fallback(rasa_response):
        #     logger.warning(f"Using LLM fallback (Rasa unavailable or low confidence)")
        #     try:
        #         response_text = ai_generator.generate_chat_response(
        #             message=message,
        #             context={'language': language, 'session_id': str(session_id), 'rasa_failed': True}
        #         )
        #         if not response_text or not response_text.strip():
        #             raise ValueError("LLM returned empty response")
        #     except Exception as llm_error:
        #         logger.error(f"LLM fallback failed: {llm_error}")
        #         response_text = ("Thank you for your message. I'm experiencing technical "
        #                          "difficulties. Please consult with our clinic staff.")
        #     response_source = "llm_fallback"
        #     buttons = []
        #     diagnosis_data = None
        #     try:
        #         predictor = get_ml_predictor()
        #         available_symptoms = predictor.get_available_symptoms()
        #         message_lower = message.lower().replace(' ', '_')
        #         extracted_symptoms = [s for s in available_symptoms if s in message_lower]
        #         if extracted_symptoms:
        #             prediction = predictor.predict(extracted_symptoms)
        #             diagnosis_data = {
        #                 'symptoms': extracted_symptoms,
        #                 'predicted_disease': prediction.get('predicted_disease'),
        #                 'confidence': prediction.get('confidence_score', 0.0),
        #                 'top_predictions': prediction.get('top_predictions', []),
        #                 'is_communicable': prediction.get('is_communicable', False),
        #                 'is_acute': prediction.get('is_acute', False),
        #                 'icd10_code': prediction.get('icd10_code', ''),
        #                 'severity': 'moderate',
        #                 'duration_days': 1
        #             }
        #     except Exception as extract_error:
        #         logger.warning(f"Could not extract symptoms: {extract_error}")
        # else:
        #     response_text = rasa_response['text']
        #     response_source = "rasa"
        #     buttons = rasa_response.get('buttons', [])
        #     diagnosis_data = rasa_response.get('custom', {}).get('diagnosis')
        #
        # record_id = None
        # if diagnosis_data and diagnosis_data.get('predicted_disease'):
        #     try:
        #         severity_map = {'mild': 1, 'moderate': 2, 'severe': 3}
        #         severity_value = diagnosis_data.get('severity', 'moderate')
        #         severity_int = severity_map.get(severity_value.lower(), 2) if isinstance(severity_value, str) else int(severity_value or 2)
        #         record = SymptomRecord.objects.create(
        #             student=request.user,
        #             symptoms=diagnosis_data.get('symptoms', []),
        #             duration_days=diagnosis_data.get('duration_days', 1),
        #             severity=severity_int,
        #             predicted_disease=diagnosis_data['predicted_disease'],
        #             confidence_score=diagnosis_data.get('confidence', 0.0),
        #             top_predictions=diagnosis_data.get('top_predictions', []),
        #             is_communicable=diagnosis_data.get('is_communicable', False),
        #             is_acute=diagnosis_data.get('is_acute', False),
        #             icd10_code=diagnosis_data.get('icd10_code', ''),
        #         )
        #         record.check_referral_criteria()
        #         record.save()
        #         FollowUp.create_from_symptom(record, days_ahead=3)
        #         record_id = str(record.id)
        #     except Exception as e:
        #         logger.error(f"Failed to create symptom record from chat: {e}")
        #
        # return Response({
        #     'response': response_text,
        #     'session_id': str(session_id),
        #     'source': response_source,
        #     'buttons': buttons,
        #     'rasa_available': rasa_service.is_available(),
        #     'record_id': record_id,
        #     'diagnosis_saved': record_id is not None,
        # })
        # ================= END RASA FLOW =================================

    except ChatSession.DoesNotExist:
        return Response(
            {'error': 'Invalid session_id'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"send_chat_message error: {e}")
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsStudent, HasDataConsent])
def generate_insights(request):
    """
    Generate top 3 health insights for current session
    POST /api/chat/insights/
    """
    session_id = request.data.get('session_id')
    symptoms = request.data.get('symptoms', [])
    disease = request.data.get('disease', '')
    
    if not session_id:
        return Response(
            {'error': 'session_id is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        session = ChatSession.objects.get(id=session_id, student=request.user)
        
        # Delete old insights for this session
        HealthInsight.objects.filter(student=request.user, session_id=session_id).delete()
        
        # Get ML predictions for context
        predictor = get_ml_predictor()
        prediction_results = predictor.predict(symptoms)
        
        # Generate new insights using LLM service
        insights_data = ai_generator.generate_health_insights(
            symptoms=symptoms,
            predictions=prediction_results,
            chat_summary=session.metadata.get('topics_discussed', '')
        )
        
        # Save top 3 insights
        insights = []
        for insight_data in insights_data[:3]:
            insight = HealthInsight.objects.create(
                student=request.user,
                session_id=session_id,
                insight_text=insight_data['text'],
                references=[],  # LLM doesn't provide references yet
                reliability_score=insight_data['reliability_score']
            )
            insights.append(insight)
        
        # Update session
        session.insights_generated_count = len(insights)
        session.save()
        
        serializer = HealthInsightSerializer(insights, many=True)
        return Response(serializer.data)
    
    except ChatSession.DoesNotExist:
        return Response(
            {'error': 'Invalid session_id'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        return Response(
            {'error': str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsStudent])
def end_chat_session(request):
    """
    End chat session and calculate duration
    POST /api/chat/end/
    """
    session_id = request.data.get('session_id')
    
    if not session_id:
        return Response(
            {'error': 'session_id is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        session = ChatSession.objects.get(id=session_id, student=request.user)
        
        if session.ended_at:
            return Response(
                {'error': 'Session already ended'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        session.ended_at = timezone.now()
        session.duration_seconds = int((session.ended_at - session.started_at).total_seconds())
        session.save()
        
        return Response({
            'message': 'Session ended',
            'duration_seconds': session.duration_seconds
        })
    
    except ChatSession.DoesNotExist:
        return Response(
            {'error': 'Invalid session_id'},
            status=status.HTTP_404_NOT_FOUND
        )


# ============================================================================
# Clinic Staff Dashboard Views
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsClinicStaff])
def clinic_dashboard(request):
    """
    Get clinic dashboard overview with statistics
    GET /api/staff/dashboard/
    """
    today = timezone.now().date()
    seven_days_ago = today - timedelta(days=7)
    thirty_days_ago = today - timedelta(days=30)
    
    # Overall statistics
    total_students = User.objects.filter(role='student').count()
    
    students_today = SymptomRecord.objects.filter(
        created_at__date=today
    ).values('student').distinct().count()
    
    students_7days = SymptomRecord.objects.filter(
        created_at__date__gte=seven_days_ago
    ).values('student').distinct().count()
    
    students_30days = SymptomRecord.objects.filter(
        created_at__date__gte=thirty_days_ago
    ).values('student').distinct().count()
    
    pending_referrals = SymptomRecord.objects.filter(
        requires_referral=True,
        referral_triggered=False
    ).count()
    
    # Department breakdown - Calculate from real data
    departments = User.objects.filter(role='student').values_list('department', flat=True).distinct()
    dept_breakdown = []
    
    for dept in departments:
        if not dept:  # Skip None/empty departments
            continue
            
        total_in_dept = User.objects.filter(role='student', department=dept).count()
        students_with_symptoms = SymptomRecord.objects.filter(
            student__department=dept,
            created_at__date__gte=thirty_days_ago
        ).values('student').distinct().count()
        
        dept_breakdown.append({
            'department': dept,
            'total_students': total_in_dept,
            'students_with_symptoms': students_with_symptoms,
            'percentage': round((students_with_symptoms / total_in_dept * 100), 1) if total_in_dept > 0 else 0
        })
    
    # Sort by students with symptoms (descending)
    dept_breakdown.sort(key=lambda x: x['students_with_symptoms'], reverse=True)
    
    # Recent symptom records
    recent_symptoms = SymptomRecord.objects.select_related('student').order_by('-created_at')[:10]
    
    # Top insight (most common disease in last 30 days)
    top_disease = SymptomRecord.objects.filter(
        created_at__date__gte=thirty_days_ago
    ).values('predicted_disease').annotate(
        count=Count('id')
    ).order_by('-count').first()
    
    top_insight = f"{top_disease['predicted_disease']} ({top_disease['count']} cases this month)" if top_disease else 'No consultations yet'
    
    # Prepare response
    data = {
        'total_students': total_students,
        'students_with_symptoms_today': students_today,
        'students_with_symptoms_7days': students_7days,
        'students_with_symptoms_30days': students_30days,
        'top_insight': top_insight,
        'department_breakdown': dept_breakdown,
        'recent_symptoms': SymptomRecordSerializer(recent_symptoms, many=True).data,
        'pending_referrals': pending_referrals
    }
    
    return Response(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsClinicStaff])
def student_directory(request):
    """
    Get filtered list of students with health records
    GET /api/staff/students/
    """
    from django.db.models import Prefetch, Avg
    from rest_framework.pagination import PageNumberPagination
    
    queryset = User.objects.filter(role='student').prefetch_related(
        'symptom_records',
        'medications',
        'follow_ups'
    ).order_by('name')
    
    # Filters
    department = request.query_params.get('department')
    search = request.query_params.get('search')
    has_symptoms = request.query_params.get('has_symptoms')
    
    if department:
        queryset = queryset.filter(department__icontains=department)
    
    if search:
        queryset = queryset.filter(
            Q(name__icontains=search) | Q(school_id__icontains=search)
        )
    
    if has_symptoms == 'true':
        queryset = queryset.filter(symptom_records__isnull=False).distinct()
    elif has_symptoms == 'false':
        queryset = queryset.filter(symptom_records__isnull=True)
    
    # Status filter
    status_filter = request.query_params.get('status')
    if status_filter == 'recent':
        seven_days_ago = timezone.now() - timedelta(days=7)
        queryset = queryset.filter(symptom_records__created_at__gte=seven_days_ago).distinct()
    elif status_filter == 'medications':
        queryset = queryset.filter(medications__is_active=True).distinct()
    elif status_filter == 'followup':
        queryset = queryset.filter(follow_ups__status__in=['pending', 'overdue']).distinct()

    # Pagination
    paginator = PageNumberPagination()
    paginator.page_size = 20
    paginator.page_size_query_param = 'page_size'
    result_page = paginator.paginate_queryset(queryset, request)

    # Build enriched student data
    students_data = []
    for student in result_page:
        # Get symptom records
        recent_symptoms = student.symptom_records.order_by('-created_at')[:5]
        last_visit = recent_symptoms.first().created_at if recent_symptoms.exists() else None
        
        # Get medications
        active_meds = student.medications.filter(is_active=True)
        
        # Calculate adherence
        med_logs = MedicationLog.objects.filter(medication__student=student)
        total_logs = med_logs.count()
        taken_logs = med_logs.filter(status='taken').count()
        adherence_rate = round((taken_logs / total_logs * 100) if total_logs > 0 else 100, 1)
        
        # Get follow-ups
        pending_followups = student.follow_ups.filter(status='pending').exists()
        
        student_data = {
            'id': student.id,
            'name': student.name,
            'school_id': student.school_id,
            'department': student.department,
            'total_visits': student.symptom_records.count(),
            'last_visit': last_visit.isoformat() if last_visit else None,
            'on_medication': active_meds.exists(),
            'medication_count': active_meds.count(),
            'adherence_rate': adherence_rate,
            'pending_followup': pending_followups,
            'recent_symptoms': recent_symptoms.exists(),
            'recent_symptom_reports': SymptomRecordSerializer(recent_symptoms, many=True).data,
            'medications': MedicationSerializer(active_meds, many=True).data
        }
        
        students_data.append(student_data)
    
    return paginator.get_paginated_response(students_data)


from rest_framework.renderers import JSONRenderer, BaseRenderer


class CSVRendererSimple(BaseRenderer):
    """
    Renderer that allows DRF to accept format=csv via query param.
    Rendering is bypassed because we return HttpResponse for CSV.
    """
    media_type = 'text/csv'
    format = 'csv'
    charset = 'utf-8'

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data


class ExcelRendererSimple(BaseRenderer):
    """
    Renderer that allows DRF to accept format=excel via query param.
    Rendering is bypassed because we return HttpResponse for Excel.
    """
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    format = 'excel'
    charset = None

    def render(self, data, accepted_media_type=None, renderer_context=None):
        return data

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsClinicStaff])
@renderer_classes([JSONRenderer, CSVRendererSimple, ExcelRendererSimple])
def export_report(request):
    """
    Export symptom data to Excel or CSV format
    GET /api/staff/export/?format=csv|excel|json&start_date=YYYY-MM-DD&end_date=YYYY-MM-DD
    
    Query Parameters:
        - format: 'json', 'csv' or 'excel' (default: csv)
        - start_date: Filter records from this date (optional)
        - end_date: Filter records until this date (optional)
        - department: Filter by department (optional)
        - disease: Filter by predicted disease (optional)
    """
    import csv
    from django.http import HttpResponse
    from datetime import datetime
    
    # Get query parameters
    export_format = request.query_params.get('format', 'csv').lower()
    start_date = request.query_params.get('start_date')
    
    # Debug: confirm view is being called
    print(f"[DEBUG] export_report called with format={export_format}")
    end_date = request.query_params.get('end_date')
    department = request.query_params.get('department')
    disease = request.query_params.get('disease')
    
    # Build queryset
    queryset = SymptomRecord.objects.select_related('student').all()
    
    if start_date:
        queryset = queryset.filter(created_at__gte=start_date)
    if end_date:
        queryset = queryset.filter(created_at__lte=end_date)
    if department:
        queryset = queryset.filter(student__department=department)
    if disease:
        queryset = queryset.filter(predicted_disease__icontains=disease)
    
    queryset = queryset.order_by('-created_at')
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # JSON format for preview
    if export_format == 'json':
        data = []
        for record in queryset:
            data.append({
                'id': str(record.id),
                'created_at': record.created_at.isoformat(),
                'student': record.student.id,
                'student_id': record.student.school_id,
                'student_name': record.student.name,
                'department': record.student.department,
                'symptoms': record.symptoms,
                'duration_days': record.duration_days,
                'severity': record.severity,
                'predicted_disease': record.predicted_disease,
                'confidence_score': record.confidence_score,
                'icd10_code': record.icd10_code,
                'is_communicable': record.is_communicable,
                'is_acute': record.is_acute,
                'requires_referral': record.requires_referral,
            })
        return Response({
            'record_count': len(data),
            'data': data,
            'generated_at': timestamp
        })
    
    if export_format == 'excel':
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Excel export requested with {queryset.count()} records")
        print(f"[DEBUG] Excel export requested with {queryset.count()} records")
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            print("[DEBUG] openpyxl imported successfully")
        except ImportError as e:
            print(f"[DEBUG] openpyxl import failed: {e}")
            return Response(
                {'error': 'Excel export requires openpyxl. Install with: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        try:
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Symptom Records"
            
            # Define headers
            headers = [
                'Date', 'Time', 'Student ID', 'Student Name', 'Department',
                'Symptoms', 'Duration (days)', 'Severity', 'Predicted Disease',
                'Confidence', 'ICD-10 Code', 'Communicable', 'Acute',
                'Requires Referral'
            ]
            
            # Style header row
            header_fill = PatternFill(start_color='006B3F', end_color='006B3F', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add data rows
            for row_num, record in enumerate(queryset, 2):
                severity_map = {1: 'Mild', 2: 'Moderate', 3: 'Severe'}
                
                ws.cell(row=row_num, column=1).value = record.created_at.strftime('%Y-%m-%d')
                ws.cell(row=row_num, column=2).value = record.created_at.strftime('%H:%M:%S')
                ws.cell(row=row_num, column=3).value = record.student.school_id
                ws.cell(row=row_num, column=4).value = record.student.name
                ws.cell(row=row_num, column=5).value = record.student.department
                ws.cell(row=row_num, column=6).value = ', '.join(record.symptoms)
                ws.cell(row=row_num, column=7).value = record.duration_days
                ws.cell(row=row_num, column=8).value = severity_map.get(record.severity, 'Unknown')
                ws.cell(row=row_num, column=9).value = record.predicted_disease
                ws.cell(row=row_num, column=10).value = f"{record.confidence_score:.1%}" if record.confidence_score else 'N/A'
                ws.cell(row=row_num, column=11).value = record.icd10_code or 'N/A'
                ws.cell(row=row_num, column=12).value = 'Yes' if record.is_communicable else 'No'
                ws.cell(row=row_num, column=13).value = 'Yes' if record.is_acute else 'No'
                ws.cell(row=row_num, column=14).value = 'Yes' if record.requires_referral else 'No'
            
            # Auto-adjust column widths
            for col_num, header in enumerate(headers, 1):
                column_letter = get_column_letter(col_num)
                max_length = len(header)
                for row in ws[column_letter]:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
            # Create HTTP response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="cpsu_health_report_{timestamp}.xlsx"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            import traceback
            print(f"[DEBUG] Excel export EXCEPTION: {e}")
            traceback.print_exc()
            return Response(
                {'error': f'Excel export failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    elif export_format == 'csv':  # CSV format
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="cpsu_health_report_{timestamp}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            'Date', 'Time', 'Student ID', 'Student Name', 'Department',
            'Symptoms', 'Duration (days)', 'Severity', 'Predicted Disease',
            'Confidence', 'ICD-10 Code', 'Communicable', 'Acute',
            'Requires Referral'
        ])
        
        # Write data rows
        severity_map = {1: 'Mild', 2: 'Moderate', 3: 'Severe'}
        
        for record in queryset:
            writer.writerow([
                record.created_at.strftime('%Y-%m-%d'),
                record.created_at.strftime('%H:%M:%S'),
                record.student.school_id,
                record.student.name,
                record.student.department,
                ', '.join(record.symptoms),
                record.duration_days,
                severity_map.get(record.severity, 'Unknown'),
                record.predicted_disease,
                f"{record.confidence_score:.1%}" if record.confidence_score else 'N/A',
                record.icd10_code or 'N/A',
                'Yes' if record.is_communicable else 'No',
                'Yes' if record.is_acute else 'No',
                'Yes' if record.requires_referral else 'No'
            ])
        
        return response
    
    else:
        return Response(
            {'error': f'Unsupported format: {export_format}. Use json, csv, or excel.'},
            status=status.HTTP_400_BAD_REQUEST
        )


# ============================================================================
# Audit Log Views (Staff Only)
# ============================================================================

class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    View audit logs (staff only, read-only)
    GET /api/audit/
    """
    queryset = AuditLog.objects.all().select_related('user').order_by('-timestamp')
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated, IsClinicStaff]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filter by action type
        action = self.request.query_params.get('action')
        if action:
            queryset = queryset.filter(action=action)
        
        # Filter by user
        school_id = self.request.query_params.get('school_id')
        if school_id:
            queryset = queryset.filter(user__school_id=school_id)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        if start_date:
            queryset = queryset.filter(timestamp__gte=start_date)
        
        return queryset


# ============================================================================
# Emergency SOS System
# ============================================================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def trigger_emergency(request):
    """
    Trigger emergency SOS alert
    POST /api/emergency/trigger/
    
    Immediately notifies all clinic staff
    """
    serializer = EmergencyTriggerSerializer(data=request.data)
    
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # Create emergency alert
    emergency = EmergencyAlert.objects.create(
        student=request.user,
        location=serializer.validated_data['location'],
        symptoms=serializer.validated_data.get('symptoms', []),
        description=serializer.validated_data.get('description', ''),
        status='active',
        priority=100  # All emergencies are critical
    )
    
    # Log the emergency
    AuditLog.objects.create(
        user=request.user,
        action='create',
        model_name='EmergencyAlert',
        object_id=str(emergency.id),
        changes={
            'location': emergency.location,
            'symptoms_count': len(emergency.symptoms)
        }
    )
    
    # TODO: Send real-time notifications to staff
    # This will be implemented with WebSockets or push notifications
    # For now, staff can poll /api/emergency/active/
    
    return Response({
        'status': 'emergency_triggered',
        'message': 'Help is on the way! Stay where you are.',
        'emergency_id': emergency.id,
        'emergency': EmergencyAlertSerializer(emergency).data
    }, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def emergency_active(request):
    """
    Get active emergencies
    GET /api/emergency/active/
    
    Students: See their own active emergencies
    Staff: See all active emergencies
    """
    if request.user.role == 'staff':
        # Staff can see all active emergencies
        emergencies = EmergencyAlert.objects.filter(
            status__in=['active', 'responding']
        ).select_related('student', 'responded_by')
    else:
        # Students see only their own
        emergencies = EmergencyAlert.objects.filter(
            student=request.user,
            status__in=['active', 'responding']
        )
    
    serializer = EmergencyAlertSerializer(emergencies, many=True)
    return Response({
        'count': emergencies.count(),
        'emergencies': serializer.data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def emergency_history(request):
    """
    Get emergency history
    GET /api/emergency/history/
    
    Students: Their own history
    Staff: All emergencies
    """
    if request.user.role == 'staff':
        emergencies = EmergencyAlert.objects.all().select_related('student', 'responded_by')
    else:
        emergencies = EmergencyAlert.objects.filter(student=request.user)
    
    # Pagination
    page_size = int(request.GET.get('page_size', 20))
    emergencies = emergencies[:page_size]
    
    serializer = EmergencyAlertSerializer(emergencies, many=True)
    return Response({
        'count': emergencies.count(),
        'emergencies': serializer.data
    })


@api_view(['PATCH'])
@permission_classes([IsClinicStaff])
def emergency_respond(request, emergency_id):
    """
    Staff responds to emergency
    PATCH /api/emergency/<id>/respond/
    """
    try:
        emergency = EmergencyAlert.objects.get(id=emergency_id)
    except EmergencyAlert.DoesNotExist:
        return Response({'error': 'Emergency not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Update status to responding if not already
    if emergency.status == 'active':
        emergency.status = 'responding'
        emergency.responded_by = request.user
        emergency.response_time = timezone.now()
        emergency.save()
        
        # Log response
        AuditLog.objects.create(
            user=request.user,
            action='update',
            model_name='EmergencyAlert',
            object_id=str(emergency.id),
            changes={
                'status': emergency.status,
                'responded_by': request.user.school_id,
                'response_time_minutes': emergency.response_time_minutes
            }
        )
    
    serializer = EmergencyAlertSerializer(emergency)
    return Response({
        'message': 'Emergency status updated to responding',
        'emergency': serializer.data
    })


@api_view(['PATCH'])
@permission_classes([IsClinicStaff])
def emergency_resolve(request, emergency_id):
    """
    Staff resolves emergency
    PATCH /api/emergency/<id>/resolve/
    """
    try:
        emergency = EmergencyAlert.objects.get(id=emergency_id)
    except EmergencyAlert.DoesNotExist:
        return Response({'error': 'Emergency not found'}, status=status.HTTP_404_NOT_FOUND)
    
    resolution_notes = request.data.get('notes', '')
    is_false_alarm = request.data.get('false_alarm', False)
    
    if is_false_alarm:
        emergency.status = 'false_alarm'
    else:
        emergency.status = 'resolved'
    
    emergency.resolved_at = timezone.now()
    emergency.resolution_notes = resolution_notes
    emergency.responded_by = request.user
    
    if not emergency.response_time:
        emergency.response_time = timezone.now()
    
    emergency.save()
    
    # Log resolution
    AuditLog.objects.create(
        user=request.user,
        action='update',
        model_name='EmergencyAlert',
        object_id=str(emergency.id),
        changes={
            'status': emergency.status,
            'resolved_at': emergency.resolved_at.isoformat() if emergency.resolved_at else None,
            'response_time_minutes': emergency.response_time_minutes
        }
    )
    
    serializer = EmergencyAlertSerializer(emergency)
    return Response({
        'message': f'Emergency {emergency.get_status_display()}',
        'emergency': serializer.data
    })


# ============================================================================
# Medication Management System
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def medication_list(request):
    """
    Get medications for current user
    GET /api/medications/
    
    Students: Their own medications
    Staff: Can filter by student_id
    """
    if request.user.role == 'staff':
        # Staff can query specific student
        student_id = request.GET.get('student_id')
        if student_id:
            medications = Medication.objects.filter(
                student__school_id=student_id
            ).select_related('student', 'prescribed_by', 'symptom_record')
        else:
            # All medications (for staff dashboard)
            medications = Medication.objects.all().select_related(
                'student', 'prescribed_by'
            )[:50]  # Limit to recent 50
    else:
        # Students see only their own
        medications = Medication.objects.filter(
            student=request.user
        ).select_related('prescribed_by', 'symptom_record')
    
    # Filter by active status
    active_only = request.GET.get('active_only', 'false').lower() == 'true'
    if active_only:
        medications = medications.filter(is_active=True)
    
    serializer = MedicationSerializer(medications, many=True)
    return Response({
        'count': medications.count(),
        'medications': serializer.data
    })


@api_view(['POST'])
@permission_classes([IsClinicStaff])
def medication_create(request):
    """
    Staff prescribes medication to student
    POST /api/medications/
    """
    serializer = MedicationCreateSerializer(data=request.data)
    
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    # Create medication
    medication = serializer.save(prescribed_by=request.user)
    
    # Auto-generate medication logs based on schedule
    from datetime import datetime, timedelta
    
    current_date = medication.start_date
    while current_date <= medication.end_date:
        for time_str in medication.schedule_times:
            # Parse time (HH:MM format)
            try:
                scheduled_time = datetime.strptime(time_str, '%H:%M').time()
                MedicationLog.objects.create(
                    medication=medication,
                    scheduled_date=current_date,
                    scheduled_time=scheduled_time,
                    status='pending'
                )
            except ValueError:
                pass  # Skip invalid time formats
        
        current_date += timedelta(days=1)
    
    # Log the prescription
    AuditLog.objects.create(
        user=request.user,
        action='create',
        model_name='Medication',
        object_id=str(medication.id),
        changes={
            'student': medication.student.school_id,
            'medication_name': medication.name,
            'duration_days': (medication.end_date - medication.start_date).days + 1,
            'schedule_times': medication.schedule_times,
        }
    )
    
    return Response({
        'message': 'Medication prescribed successfully',
        'medication': MedicationSerializer(medication).data
    }, status=status.HTTP_201_CREATED)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def medication_detail(request, medication_id):
    """
    Get medication details
    GET /api/medications/<id>/
    """
    try:
        medication = Medication.objects.select_related(
            'student', 'prescribed_by', 'symptom_record'
        ).prefetch_related('logs').get(id=medication_id)
    except Medication.DoesNotExist:
        return Response({'error': 'Medication not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Permission check
    if request.user.role != 'staff' and medication.student != request.user:
        return Response(
            {'error': 'You do not have permission to view this medication'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    serializer = MedicationSerializer(medication)
    return Response(serializer.data)


@api_view(['PATCH'])
@permission_classes([IsClinicStaff])
def medication_update(request, medication_id):
    """
    Update medication (staff only)
    PATCH /api/medications/<id>/
    """
    try:
        medication = Medication.objects.get(id=medication_id)
    except Medication.DoesNotExist:
        return Response({'error': 'Medication not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Update fields
    allowed_fields = ['dosage', 'frequency', 'instructions', 'is_active', 'end_date']
    for field in allowed_fields:
        if field in request.data:
            setattr(medication, field, request.data[field])
    
    medication.save()
    
    return Response({
        'message': 'Medication updated',
        'medication': MedicationSerializer(medication).data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def medication_logs_today(request):
    """
    Get today's medication schedule for student
    GET /api/medications/logs/today/
    """
    today = timezone.now().date()
    
    if request.user.role == 'student':
        # Get student's medications
        medications = Medication.objects.filter(
            student=request.user,
            is_active=True
        )
        
        logs = MedicationLog.objects.filter(
            medication__in=medications,
            scheduled_date=today
        ).select_related('medication').order_by('scheduled_time')
        
    else:
        # Staff sees all for the day
        logs = MedicationLog.objects.filter(
            scheduled_date=today
        ).select_related('medication', 'medication__student')[:100]
    
    serializer = MedicationLogSerializer(logs, many=True)
    return Response({
        'date': today,
        'count': logs.count(),
        'logs': serializer.data
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def medication_log_mark_taken(request, log_id):
    """
    Mark medication dose as taken
    POST /api/medications/logs/<id>/taken/
    """
    try:
        log = MedicationLog.objects.select_related('medication').get(id=log_id)
    except MedicationLog.DoesNotExist:
        return Response({'error': 'Log not found'}, status=status.HTTP_404_NOT_FOUND)
    
    # Permission check
    if request.user.role != 'staff' and log.medication.student != request.user:
        return Response(
            {'error': 'Permission denied'},
            status=status.HTTP_403_FORBIDDEN
        )
    
    notes = request.data.get('notes', '')
    log.mark_as_taken(notes=notes)
    
    return Response({
        'message': 'Marked as taken',
        'log': MedicationLogSerializer(log).data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def medication_adherence(request):
    """
    Get medication adherence statistics
    GET /api/medications/adherence/
    """
    if request.user.role == 'student':
        medications = Medication.objects.filter(student=request.user, is_active=True)
    else:
        student_id = request.GET.get('student_id')
        if student_id:
            medications = Medication.objects.filter(
                student__school_id=student_id,
                is_active=True
            )
        else:
            return Response({'error': 'student_id required for staff'}, status=400)
    
    stats = []
    for med in medications:
        total_logs = med.logs.exclude(status='pending').count()
        if total_logs > 0:
            taken_count = med.logs.filter(status='taken').count()
            missed_count = med.logs.filter(status='missed').count()
            adherence_rate = (taken_count / total_logs) * 100
            
            stats.append({
                'medication_id': str(med.id),
                'medication_name': med.name,
                'total_doses': total_logs,
                'taken': taken_count,
                'missed': missed_count,
                'adherence_rate': round(adherence_rate, 1),
                'days_remaining': med.days_remaining
            })
    
    # Overall adherence
    total_all = sum(s['total_doses'] for s in stats)
    taken_all = sum(s['taken'] for s in stats)
    overall_rate = (taken_all / total_all * 100) if total_all > 0 else 0
    
    return Response({
        'overall_adherence_rate': round(overall_rate, 1),
        'medications': stats
    })


# ============================================================================
# Follow-Up Views
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def followup_list(request):
    """
    Get follow-ups for current user (students) or all/specific student (staff)
    GET /api/followups/
    Query params: status, student_id (staff only)
    """
    if request.user.role == 'student':
        followups = FollowUp.objects.filter(student=request.user)
    else:
        # Staff can see all or filter by student
        student_id = request.GET.get('student_id')
        if student_id:
            followups = FollowUp.objects.filter(student__school_id=student_id)
        else:
            followups = FollowUp.objects.all()
    
    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter:
        followups = followups.filter(status=status_filter)
    
    # Auto-check and update overdue status
    for followup in followups:
        followup.check_overdue()
    
    serializer = FollowUpSerializer(followups, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def followup_pending(request):
    """
    Get pending follow-ups (including overdue) for current user
    GET /api/followups/pending/
    """
    followups = FollowUp.objects.filter(
        student=request.user,
        status__in=['pending', 'overdue']
    ).order_by('scheduled_date')
    
    # Check for overdue
    for followup in followups:
        followup.check_overdue()
    
    serializer = FollowUpSerializer(followups, many=True)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsStudent])
def followup_respond(request, pk):
    """
    Submit response to follow-up
    POST /api/followups/<id>/respond/
    """
    try:
        followup = FollowUp.objects.get(pk=pk, student=request.user)
    except FollowUp.DoesNotExist:
        return Response({'error': 'Follow-up not found'}, status=404)
    
    if followup.status == 'completed':
        return Response({'error': 'Follow-up already completed'}, status=400)
    
    serializer = FollowUpResponseSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=400)
    
    # Update follow-up with response
    followup.outcome = serializer.validated_data['outcome']
    followup.notes = serializer.validated_data.get('notes', '')
    followup.still_experiencing_symptoms = serializer.validated_data['still_experiencing_symptoms']
    followup.new_symptoms = serializer.validated_data.get('new_symptoms', [])
    followup.response_date = timezone.now()
    followup.status = 'completed'
    followup.save()
    
    # Auto-flag for appointment if condition worsened
    if followup.outcome == 'worse':
        followup.requires_appointment = True
        followup.save()
    
    return Response({
        'message': 'Follow-up response submitted',
        'followup': FollowUpSerializer(followup).data
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated, IsClinicStaff])
def followup_review(request, pk):
    """
    Staff review of follow-up response
    POST /api/followups/<id>/review/
    """
    try:
        followup = FollowUp.objects.get(pk=pk)
    except FollowUp.DoesNotExist:
        return Response({'error': 'Follow-up not found'}, status=404)
    
    review_notes = request.data.get('review_notes', '')
    requires_appointment = request.data.get('requires_appointment', False)
    
    followup.reviewed_by = request.user
    followup.review_notes = review_notes
    followup.requires_appointment = requires_appointment
    followup.status = 'reviewed'
    followup.response_date = timezone.now()
    followup.save()
    
    return Response({
        'message': 'Follow-up reviewed',
        'followup': FollowUpSerializer(followup).data
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated, IsClinicStaff])
def followup_needs_review(request):
    """
    Get follow-ups that need staff review
    GET /api/followups/needs-review/
    """
    # Return all follow-ups; frontend filters for counts
    followups = FollowUp.objects.select_related('student', 'reviewed_by', 'symptom_record')\
        .order_by('-scheduled_date')

    # Enrich with student data
    data = []
    for followup in followups:
        followup_data = FollowUpSerializer(followup).data
        followup_data['student_name'] = followup.student.name
        followup_data['student_school_id'] = followup.student.school_id
        followup_data['student_department'] = followup.student.department
        followup_data['reviewed_by_name'] = followup.reviewed_by.name if followup.reviewed_by else None
        data.append(followup_data)

    return Response(data)


# ============================================================================
# Analytics Views (Real Data)
# ============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated, IsClinicStaff])
def staff_analytics(request):
    """
    Get comprehensive analytics data for charts
    GET /api/staff/analytics/?period=7d
    
    Periods: 7d, 30d, 90d, 1y
    """
    period = request.query_params.get('period', '30d')
    
    # Calculate date range
    today = timezone.now().date()
    if period == '7d':
        start_date = today - timedelta(days=7)
    elif period == '30d':
        start_date = today - timedelta(days=30)
    elif period == '90d':
        start_date = today - timedelta(days=90)
    elif period == '1y':
        start_date = today - timedelta(days=365)
    else:
        start_date = today - timedelta(days=30)
    
    # Summary stats
    total_consultations = SymptomRecord.objects.filter(created_at__date__gte=start_date).count()
    unique_patients = SymptomRecord.objects.filter(created_at__date__gte=start_date).values('student').distinct().count()
    emergency_alerts = EmergencyAlert.objects.filter(created_at__date__gte=start_date).count()
    prescriptions = Medication.objects.filter(created_at__date__gte=start_date).count()
    
    # Top 10 diagnosed conditions
    top_conditions = SymptomRecord.objects.filter(
        created_at__date__gte=start_date
    ).values('predicted_disease').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Consultation trends (daily counts)
    consultation_trends = []
    current_date = start_date
    while current_date <= today:
        count = SymptomRecord.objects.filter(created_at__date=current_date).count()
        consultation_trends.append({
            'date': current_date.isoformat(),
            'count': count
        })
        current_date += timedelta(days=1)
    
    # Consultations by department
    dept_breakdown = SymptomRecord.objects.filter(
        created_at__date__gte=start_date
    ).values('student__department').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Symptom severity distribution (using actual severity field: 1=Mild, 2=Moderate, 3=Severe)
    severity_distribution = {
        'mild': SymptomRecord.objects.filter(created_at__date__gte=start_date, severity=1).count(),
        'moderate': SymptomRecord.objects.filter(created_at__date__gte=start_date, severity=2).count(),
        'severe': SymptomRecord.objects.filter(created_at__date__gte=start_date, severity=3).count(),
    }
    
    # Most common symptoms
    from collections import Counter
    
    symptom_records = SymptomRecord.objects.filter(created_at__date__gte=start_date)
    all_symptoms = []
    for record in symptom_records:
        if record.symptoms:  # Changed from symptoms_reported to symptoms
            all_symptoms.extend(record.symptoms)
    
    symptom_counter = Counter(all_symptoms)
    common_symptoms = [
        {
            'symptom': symptom,
            'count': count,
            'percentage': round((count / len(all_symptoms) * 100) if all_symptoms else 0, 1)
        }
        for symptom, count in symptom_counter.most_common(10)
    ]
    
    data = {
        'period': period,
        'summary': {
            'total_consultations': total_consultations,
            'unique_patients': unique_patients,
            'emergency_alerts': emergency_alerts,
            'prescriptions': prescriptions
        },
        'top_conditions': list(top_conditions),
        'consultation_trends': consultation_trends,
        'department_breakdown': list(dept_breakdown),
        'severity_distribution': severity_distribution,
        'common_symptoms': common_symptoms
    }
    
    return Response(data)

# ============================================================================
# Messaging System
# ============================================================================

class MessageViewSet(viewsets.ModelViewSet):
    """
    CRUD for direct messages
    """
    serializer_class = MessageSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        return Message.objects.filter(
            Q(sender=user) | Q(recipient=user)
        ).select_related('sender', 'recipient')

    def create(self, request, *args, **kwargs):
        # Allow looking up recipient by school_id or UUID
        if 'recipient' in request.data:
            recipient_data = request.data['recipient']
            if isinstance(recipient_data, str) and not recipient_data.isdigit():
                # First, check if it's a valid UUID (for reply-by-PK)
                try:
                    import uuid as uuid_mod
                    uuid_mod.UUID(str(recipient_data))
                    # It's a valid UUID - let the default serializer handle it
                    return super().create(request, *args, **kwargs)
                except (ValueError, AttributeError):
                    pass

                # Otherwise, try to find user by school_id
                try:
                    recipient_user = User.objects.get(school_id=recipient_data)
                    data = request.data.copy()
                    data['recipient'] = recipient_user.id

                    serializer = self.get_serializer(data=data)
                    serializer.is_valid(raise_exception=True)
                    self.perform_create(serializer)
                    headers = self.get_success_headers(serializer.data)
                    return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

                except User.DoesNotExist:
                    return Response(
                        {'recipient': [f'User with school_id {recipient_data} not found.']},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                except Exception as e:
                    logger.error(f"Error resolving recipient: {e}")

        return super().create(request, *args, **kwargs)

    @action(detail=False, methods=['get'])
    def user_search(self, request):
        """Search for a user by school_id to compose a message"""
        school_id = request.query_params.get('school_id', '').strip()
        if not school_id:
            return Response({'error': 'school_id parameter required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            user = User.objects.get(school_id=school_id)
            return Response({
                'id': str(user.id),
                'name': user.name,
                'school_id': user.school_id,
                'role': user.role,
            })
        except User.DoesNotExist:
            return Response({'error': 'User not found'}, status=status.HTTP_404_NOT_FOUND)

    def perform_create(self, serializer):
        try:
            serializer.save(sender=self.request.user)
        except IntegrityError:
            logger.error(f"Database integrity error while creating message. User: {self.request.user.school_id}")
            raise serializers.ValidationError({"detail": "Database integrity error. Please check your input."})
        except Exception as e:
            logger.exception(f"Unexpected error creating message for user {self.request.user.school_id}: {str(e)}")
            raise APIException("An internal error occurred while sending the message. Please contact support.")

    @action(detail=True, methods=['patch'])
    def mark_read(self, request, pk=None):
        message = self.get_object()
        if message.recipient != request.user:
             return Response({'error': 'Not authorized'}, status=status.HTTP_403_FORBIDDEN)

        message.is_read = True
        message.save()
        return Response({'status': 'marked as read'})


# ============================================================================
# Appointment System
# ============================================================================

class AppointmentViewSet(viewsets.ModelViewSet):
    """
    CRUD for appointments
    """
    serializer_class = AppointmentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role == 'staff':
            return Appointment.objects.all().select_related('student', 'staff')
        return Appointment.objects.filter(student=user).select_related('student', 'staff')

    def perform_create(self, serializer):
        # If student creates, set student=user. If staff creates, they must specify student.
        if self.request.user.role == 'student':
            serializer.save(student=self.request.user)
        else:
            serializer.save()

    @action(detail=True, methods=['patch'])
    def confirm(self, request, pk=None):
        if request.user.role != 'staff':
            return Response({'error': 'Only staff can confirm appointments'}, status=status.HTTP_403_FORBIDDEN)
        appointment = self.get_object()
        appointment.status = 'confirmed'
        appointment.staff = request.user
        appointment.save()
        return Response(AppointmentSerializer(appointment).data)
